from pathlib import Path
from shutil import copyfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT_OUTPUT = Path(r"D:\quan ly cong viec\mau-du-lieu-day-len-web.xlsx")
PUBLIC_OUTPUT = Path(r"D:\quan ly cong viec\app\public\mau-tong-hop.xlsx")

TASKS = [
    "Số thu",
    "Kê khai thuế",
    "Quản lý rủi ro HKD",
    "Kiểm tra HKD",
    "Rà soát TMĐT",
    "Hỗ trợ hóa đơn điện tử",
    "Chuyển đổi lên doanh nghiệp",
    "Nộp thuế điện tử",
    "Nợ thuế",
    "Cưỡng chế xuất nhập cảnh",
    "Cưỡng chế tài khoản hóa đơn",
    "Hệ số K",
    "Thủ tục hành chính",
]

OFFICERS = [
    "Ngô Mai Trang",
    "Công Tiến Tùng",
    "Nguyễn Thị Lệ",
    "Nguyễn Văn Tuấn",
    "Nguyễn Tùng Dương",
    "Nguyễn Thị Hương Hà",
    "Nguyễn Đức Mạnh",
    "Nguyễn Thị Thu Trà",
    "Nguyễn Thị Thùy Liên",
    "Hoàng Thế Vương",
    "Nguyễn Thị Thu Hoài",
    "Trần Thanh Thư",
    "Nguyễn Việt Toàn",
    "Nguyễn Kim Ngân",
]

OLD_AREAS = [
    "xã An Khánh (hết hiệu lực)",
    "xã Vân Côn (hết hiệu lực)",
    "xã Vân Canh (hết hiệu lực)",
    "xã Sơn Đồng (hết hiệu lực)",
    "xã Song Phương (hết hiệu lực)",
    "xã Tiền Yên (hết hiệu lực)",
    "xã Lại Yên (hết hiệu lực)",
    "xã La Phù (hết hiệu lực)",
    "xã Cát Quế (hết hiệu lực)",
    "xã Đắc Sở (hết hiệu lực)",
    "thị trấn Trạm Trôi (hết hiệu lực)",
    "xã Yên Sở (hết hiệu lực)",
    "xã Di Trạch (hết hiệu lực)",
    "xã Dương Liễu (hết hiệu lực)",
    "xã Đức Thượng (hết hiệu lực)",
    "xã Đức Giang (hết hiệu lực)",
    "xã Kim Chung (hết hiệu lực)",
    "xã Minh Khai (hết hiệu lực)",
    "xã An Thượng (hết hiệu lực)",
    "xã Đông La (hết hiệu lực)",
]

NEW_AREAS = ["xã Hoài Đức", "xã Sơn Đồng", "xã An Khánh", "xã Dương Hòa"]

TEAMS = [
    "Tổ Quản lý hỗ trợ cá nhân hộ kinh doanh số 1",
    "Tổ Quản lý hỗ trợ cá nhân hộ kinh doanh số 2",
]

HEADERS = [
    "STT",
    "Tên nhiệm vụ",
    "Cán bộ",
    "Địa bàn xã cũ",
    "Địa bàn xã mới",
    "Tổ quản lý",
    "Mã số thuế",
    "CCCD",
    "Phải thực hiện",
    "Đã thực hiện",
    "Thời hạn",
    "Ghi chú",
]


def style_range(ws, max_row: int, max_col: int, border: Border) -> None:
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def create_template() -> None:
    PUBLIC_OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Du lieu"
    ws.append(HEADERS)
    for index in range(1, 301):
        ws.append([index, "", "", "", "", "", "", "", "", "", "", ""])

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    subtle_fill = PatternFill("solid", fgColor="F8FAFC")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    style_range(ws, 301, len(HEADERS), border)
    for row in range(2, 302):
        ws[f"A{row}"].fill = subtle_fill
        ws[f"A{row}"].font = Font(color="475569", bold=True)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"I{row}"].number_format = "#,##0"
        ws[f"J{row}"].number_format = "#,##0"
        ws[f"K{row}"].number_format = "yyyy-mm-dd"

    widths = {
        "A": 8,
        "B": 30,
        "C": 26,
        "D": 34,
        "E": 22,
        "F": 46,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 34,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:L301"
    ws.sheet_view.showGridLines = False

    list_ws = wb.create_sheet("Danh muc")
    list_ws.append(["Tên nhiệm vụ", "Cán bộ", "Địa bàn xã cũ", "Địa bàn xã mới", "Tổ quản lý"])
    max_len = max(len(TASKS), len(OFFICERS), len(OLD_AREAS), len(NEW_AREAS), len(TEAMS))
    for index in range(max_len):
        list_ws.append(
            [
                TASKS[index] if index < len(TASKS) else "",
                OFFICERS[index] if index < len(OFFICERS) else "",
                OLD_AREAS[index] if index < len(OLD_AREAS) else "",
                NEW_AREAS[index] if index < len(NEW_AREAS) else "",
                TEAMS[index] if index < len(TEAMS) else "",
            ]
        )
    for cell in list_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    style_range(list_ws, max_len + 1, 5, border)
    for column, width in {"A": 32, "B": 26, "C": 34, "D": 22, "E": 46}.items():
        list_ws.column_dimensions[column].width = width
    list_ws.freeze_panes = "A2"
    list_ws.sheet_view.showGridLines = False

    help_ws = wb.create_sheet("Huong dan")
    help_rows = [
        ["Mục", "Hướng dẫn"],
        [
            "Cách dùng",
            "Nhập dữ liệu vào sheet Du lieu, sau đó upload file bằng nút mũi tên lên hoặc đưa file lên Google Drive/Sheets rồi dán link Drive.",
        ],
        ["Tên nhiệm vụ", "Bắt buộc. Chọn trong danh mục để web ghép đúng chuyên đề."],
        ["Cán bộ", "Bắt buộc. Chọn đúng tên cán bộ đang có trên web."],
        [
            "Địa bàn xã cũ",
            "Có dropdown chọn; tất cả xã cũ đều kèm trạng thái hết hiệu lực. Cột này để quản lý, web hiện không bắt buộc để cập nhật chỉ tiêu.",
        ],
        ["Địa bàn xã mới", "Có dropdown chọn: xã Hoài Đức, xã Sơn Đồng, xã An Khánh, xã Dương Hòa."],
        ["Tổ quản lý", "Nên chọn đúng tổ. Nếu để trống, web vẫn có thể nhận theo tên cán bộ nếu tìm thấy."],
        ["Mã số thuế / CCCD", "Không bắt buộc, dùng để lưu thông tin tra cứu nếu có."],
        ["Phải thực hiện", "Nhập số chỉ tiêu giao, số nguyên không âm."],
        ["Đã thực hiện", "Nhập số đã rà soát/đã đạt, số nguyên không âm."],
        ["Thời hạn", "Nên nhập dạng yyyy-mm-dd, ví dụ 2026-10-06."],
        ["Lưu ý", "Không đổi tên các cột chính: Tên nhiệm vụ, Cán bộ, Tổ quản lý, Phải thực hiện, Đã thực hiện, Thời hạn."],
    ]
    for row in help_rows:
        help_ws.append(row)
    for cell in help_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    style_range(help_ws, len(help_rows), 2, border)
    help_ws.column_dimensions["A"].width = 22
    help_ws.column_dimensions["B"].width = 96
    help_ws.sheet_view.showGridLines = False

    validations = [
        (DataValidation(type="list", formula1="='Danh muc'!$A$2:$A$14", allow_blank=True), "B2:B301", "Chọn tên nhiệm vụ trong danh mục."),
        (DataValidation(type="list", formula1="='Danh muc'!$B$2:$B$15", allow_blank=True), "C2:C301", "Chọn tên cán bộ trong danh mục."),
        (DataValidation(type="list", formula1="='Danh muc'!$C$2:$C$21", allow_blank=True), "D2:D301", "Chọn địa bàn xã cũ trong danh mục."),
        (DataValidation(type="list", formula1="='Danh muc'!$D$2:$D$5", allow_blank=True), "E2:E301", "Chọn địa bàn xã mới trong danh mục."),
        (DataValidation(type="list", formula1="='Danh muc'!$E$2:$E$3", allow_blank=True), "F2:F301", "Chọn tổ quản lý trong danh mục."),
        (DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True), "I2:J301", "Chỉ nhập số nguyên không âm."),
        (DataValidation(type="date", operator="greaterThanOrEqual", formula1="DATE(2024,1,1)", allow_blank=True), "K2:K301", "Nhập ngày hợp lệ, nên dùng yyyy-mm-dd."),
    ]
    for validation, target, message in validations:
        validation.errorTitle = "Dữ liệu không hợp lệ"
        validation.error = message
        ws.add_data_validation(validation)
        validation.add(target)

    table = Table(displayName="BangDuLieuWeb", ref="A1:L301")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(ROOT_OUTPUT)
    copyfile(ROOT_OUTPUT, PUBLIC_OUTPUT)

    check = load_workbook(ROOT_OUTPUT, data_only=False)
    assert check.sheetnames == ["Du lieu", "Danh muc", "Huong dan"]
    check_ws = check["Du lieu"]
    assert [check_ws.cell(1, column).value for column in range(1, 13)] == HEADERS
    assert check["Danh muc"]["A2"].value == "Số thu"
    assert check["Danh muc"]["C21"].value == "xã Đông La (hết hiệu lực)"
    assert len(check_ws.data_validations.dataValidation) == 7


if __name__ == "__main__":
    create_template()
    print(ROOT_OUTPUT)
    print(PUBLIC_OUTPUT)
