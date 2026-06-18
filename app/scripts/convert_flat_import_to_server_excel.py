from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
import sys
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OFFICERS = [
    "Ngô Mai Trang",
    "Công Tiến Tùng",
    "Nguyễn Thị Lệ",
    "Nguyễn Văn Tuấn",
    "Nguyễn Tùng Dương",
    "Nguyễn Thị Hương Hà",
    "Nguyễn Đức Mạnh",
    "Nguyễn Thị Thu Trà",
    "Nguyễn Thị Thùy Liên",
    "Hoàng Thế Vương",
    "Nguyễn Thị Thu Hoài",
    "Trần Thanh Thư",
    "Nguyễn Việt Toàn",
    "Nguyễn Kim Ngân",
]

OFFICER_ALIASES = {
    "Nguyễn Thùy Liên": "Nguyễn Thị Thùy Liên",
}

TASKS = [
    "Số thu",
    "Kê khai thuế",
    "Quản lý rủi ro HKD",
    "Kiểm tra HKD",
    "Rà soát TMĐT",
    "Hỗ trợ hóa đơn điện tử",
    "Chuyển đổi lên doanh nghiệp",
    "Nộp thuế điện tử",
    "Nợ thuế",
    "Cưỡng chế xuất nhập cảnh",
    "Cưỡng chế tài khoản hóa đơn",
    "Hệ số K",
    "Thủ tục hành chính",
]

TASK_START_COLUMNS = {
    "Số thu": 7,
    "Kê khai thuế": 10,
    "Quản lý rủi ro HKD": 13,
    "Kiểm tra HKD": 16,
    "Rà soát TMĐT": 19,
    "Hỗ trợ hóa đơn điện tử": 22,
    "Chuyển đổi lên doanh nghiệp": 25,
    "Nộp thuế điện tử": 28,
    "Nợ thuế": 31,
    "Cưỡng chế xuất nhập cảnh": 34,
    "Cưỡng chế tài khoản hóa đơn": 37,
    "Hệ số K": 40,
    "Thủ tục hành chính": 43,
}


def normalize(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    return re.sub(r"\s+", " ", text).lower().strip()


def number(value: object) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    text = re.sub(r"[^\d.-]", "", text)
    return float(text) if text else 0


def date_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = "" if value is None else str(value).strip()
    if not text:
        return "2026-10-06"
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10] or "2026-10-06"


def find_column(headers: list[object], aliases: list[str]) -> int:
    normalized_aliases = {normalize(alias) for alias in aliases}
    for index, header in enumerate(headers, start=1):
        if normalize(header) in normalized_aliases:
            return index
    raise ValueError(f"Không tìm thấy cột: {aliases[0]}")


def split_number(total: float, parts: int) -> list[float]:
    total = round(total)
    if parts <= 1:
        return [total]
    base = total // parts
    remainder = total - base * parts
    return [base + (1 if index < remainder else 0) for index in range(parts)]


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: convert_flat_import_to_server_excel.py input.xlsx output.xlsx")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    source = load_workbook(input_path, data_only=True)
    ws = source["Du lieu"] if "Du lieu" in source.sheetnames else source[source.sheetnames[0]]
    headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]

    task_col = find_column(headers, ["Tên nhiệm vụ", "Nhiệm vụ", "Công việc"])
    officer_col = find_column(headers, ["Cán bộ", "Người thực hiện", "Người được giao"])
    assigned_col = find_column(headers, ["Phải thực hiện", "Tổng phải thực hiện", "Chỉ tiêu"])
    completed_col = find_column(headers, ["Đã thực hiện", "Số đã thực hiện", "Thực hiện"])
    try:
        deadline_col = find_column(headers, ["Thời hạn", "Deadline", "Hạn xử lý"])
    except ValueError:
        deadline_col = 0

    officer_lookup = {normalize(name): name for name in OFFICERS}
    officer_lookup.update({normalize(alias): canonical for alias, canonical in OFFICER_ALIASES.items()})
    task_lookup = {normalize(title): title for title in TASKS}
    data: dict[tuple[str, str], dict[str, object]] = defaultdict(
        lambda: {"assigned": 0, "completed": 0, "deadline": "2026-10-06"}
    )

    imported_rows = 0
    expanded_rows = 0
    skipped_rows = 0
    for row_index in range(2, ws.max_row + 1):
        raw_task = ws.cell(row_index, task_col).value
        raw_officer = ws.cell(row_index, officer_col).value
        if not raw_task or not raw_officer:
            continue

        task = task_lookup.get(normalize(raw_task))
        officer_names = [name.strip() for name in re.split(r"\s*[-–—]\s*", str(raw_officer)) if name.strip()]
        officers = [officer_lookup[normalize(name)] for name in officer_names if normalize(name) in officer_lookup]
        if not task or not officers:
            skipped_rows += 1
            continue

        raw_assigned = number(ws.cell(row_index, assigned_col).value)
        raw_completed = number(ws.cell(row_index, completed_col).value)
        assigned_parts = split_number(raw_assigned, len(officers))
        completed_parts = split_number(raw_completed, len(officers))
        for index, officer in enumerate(officers):
            key = (officer, task)
            data[key]["assigned"] = round(float(data[key]["assigned"]) + assigned_parts[index], 4)
            data[key]["completed"] = round(float(data[key]["completed"]) + completed_parts[index], 4)
            if deadline_col:
                data[key]["deadline"] = date_text(ws.cell(row_index, deadline_col).value)
        imported_rows += 1
        expanded_rows += len(officers)

    wb = Workbook()
    out = wb.active
    out.title = "Du lieu"
    out.cell(1, 1, "Tên cán bộ")
    for task in TASKS:
        col = TASK_START_COLUMNS[task]
        out.cell(1, col, task)
        out.cell(2, col, "Số phải thực hiện")
        out.cell(2, col + 1, "Số đã thực hiện")
        out.cell(2, col + 2, "Thời hạn")

    for row_offset, officer in enumerate(OFFICERS, start=3):
        out.cell(row_offset, 1, officer)
        for task in TASKS:
            col = TASK_START_COLUMNS[task]
            values = data[(officer, task)]
            out.cell(row_offset, col, values["assigned"])
            out.cell(row_offset, col + 1, values["completed"])
            out.cell(row_offset, col + 2, values["deadline"])

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    sub_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in out.iter_rows(min_row=1, max_row=16, min_col=1, max_col=45):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in out[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in out[2]:
        cell.fill = sub_fill
        cell.font = Font(color="1E3A8A", bold=True)
    out.column_dimensions["A"].width = 28
    for column in range(7, 46):
        out.column_dimensions[out.cell(1, column).column_letter].width = 16
    out.freeze_panes = "G3"
    out.sheet_view.showGridLines = False

    wb.save(output_path)
    print(f"imported_rows={imported_rows}")
    print(f"expanded_rows={expanded_rows}")
    print(f"skipped_rows={skipped_rows}")
    print(output_path)


if __name__ == "__main__":
    main()
